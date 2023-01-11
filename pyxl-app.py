from openpyxl import load_workbook, Workbook

path_to_data_1 = "C:\\Users\\bryndell.torio\\OneDrive - Integrated Micro-Electronics Inc\\Design\\Training\\pyxl_training\\Data\\bills_of_materials.xlsx"
path_to_data_2 = "C:\\Users\\bryndell.torio\\OneDrive - Integrated Micro-Electronics Inc\\Design\\Training\\pyxl_training\\Data\\test_data.xlsx"

wb = load_workbook(filename = path_to_data_1)
ws = wb.active

workbook_max_row = ws.max_row
workbook_max_column = ws.max_column

# Sample data to test the function definition.
sample_list = ['C606','C1514','C1801']
cell = ws.cell(row = 17, column = 3)

# Converts the cell object into string the split it using the comma ',' delimiter.
tmpcell = str(cell.value).split(',')


# Function definition that takes a 2 list then returns a list of similar items.
def compare_list(original_workbook,t_workbook):
    comparison_results = []
    for i in range(len(original_workbook)):
        for j in range(len(t_workbook)):
            if (original_workbook[i] == t_workbook[j]):
                comparison_results.append(original_workbook[i])
    return comparison_results


print(compare_list(tmpcell,sample_list))
