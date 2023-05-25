import customtkinter
import os
from PIL import Image


#imports to copy paste from initial BOM to master BOM.
from pathlib import Path
from openpyxl import load_workbook, Workbook
import pandas as pd


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Integrated Micro-Electronics Inc. - Design and Development")
        self.geometry("700x450")

        # set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # load images with light and dark mode image
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Images")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "imi_logo.png")), size=(26, 26))
        self.circuit_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "circuit_board_2.png")), size=(500, 150))
        self.excel_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "excel_icon_light.png")), size=(20, 20))
        self.file_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "file_icon_light.png")), size=(20, 20))
        self.imi_box_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "imi_box_logo.png")), size=(20, 20))
        self.information_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "information_icon_light.png")), size=(20, 20))
        self.home_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "home_light.png")), size=(20, 20))

        # create navigation frame
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text="  BOM Generator", image=self.imi_box_image,
                                                             compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Information",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.information_image, anchor="w", command=self.frame_2_button_event)
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["Altium", "Cadence", "PADS"],
                                                                command=self.select_CAD_event)
        self.appearance_mode_menu.grid(row=6, column=0, padx=20, pady=20, sticky="s")

        # create home frame
        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(0, weight=1)

        self.home_frame_large_image_label = customtkinter.CTkLabel(self.home_frame, text="", image=self.circuit_image)
        self.home_frame_large_image_label.grid(row=0, column=0, padx=30, pady=10)

        self.home_frame_button_1 = customtkinter.CTkButton(self.home_frame, text="Choose file", image=self.excel_image, compound="left", command=self.copy_files_to_Master)
        self.home_frame_button_1.grid(row=1, column=0, padx=20, pady=10)
        self.home_frame_button_2 = customtkinter.CTkButton(self.home_frame, text="logfile", image=self.file_image, compound="left")
        self.home_frame_button_2.grid(row=2, column=0, padx=20, pady=10)

        # create second frame
        self.second_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # create third frame
        self.third_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        # select default frame
        self.select_frame_by_name("home")

    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")

        # show selected frame
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.second_frame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def frame_3_button_event(self):
        self.select_frame_by_name("frame_3")

    def select_CAD_event(self, new_appearance_mode):
        # insert code to select the right command
        if new_appearance_mode == "Altium":
            print("Altium is selected.")
        elif new_appearance_mode == "Cadence":
            print("Cadence is selected.")
        elif new_appearance_mode == "PADS":
            print("PADS is selected.")

    ### The function below have UnicodeError Bug. Needs to be fixed for converting CSV to Excel functionality.
    # def convert_csv_to_excel(self):
    #     SOURCE_DIR = os.getcwd()
    #     DATA_LOCATION = "\\tmp\\"
    #
    #     path_to_data = customtkinter.filedialog.askopenfilename(
    #             initialdir=SOURCE_DIR,
    #             title="Select a file",
    #             filetypes=(("CSV files", "*.csv"),
    #                        ("Excel files", "*.xlsx"),
    #                        ("all files", "*.*")))
    #     fileNameWithOutExtension = Path(path_to_data).stem
    #     excelFileNameFullPath = SOURCE_DIR + DATA_LOCATION + str(fileNameWithOutExtension) + ".xlsx"
    #     read_file = pd.read_csv(path_to_data, engine='python')
    #     print(path_to_data)
    #     read_file.to_excel(excelFileNameFullPath,
    #                        index=None, 
    #                        header=True) 

    def copy_files_to_Master(self):
        self.convert_csv_to_excel()
        # COMPONENT_START_INDEX = 15 - 1 #The indexing starts at 0.
        # dict_item = {}
        # wb = load_workbook(filename=SAMPLE_DATA)
        # ws = wb.active
        # indexLastElement = ws.max_row

        # totalNumberOfItems = indexLastElement - COMPONENT_START_INDEX

        # Name_ColumnTitle = ws["A12"].value
        # lastElement = "A" + str(indexLastElement)
        # firstElement = "A" + str(COMPONENT_START_INDEX + 1)
        # column_list = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
        # for i in column_list:
        #     if column_list is not None:
        #         print(ws[f"{i}12"].value)
        #     else:
        #         return
        # print(type(column_list[1]))

        # for i in column_list:
        #     placeHolder = str(column_list[i]) + "12"
        #     print(ws[placeHolder].value)

        # for i in column_list:
            # Name_ColumnTitle = ws[column_list[i] + "12"].value
            # lastElement = column_list[i] + str(indexLastElement)
            # firstElement = column_list[i] + str(COMPONENT_START_INDEX + 1)
            # rng = wb[ws.title][firstElement:lastElement]
            # rng_values = []
            # for cells in rng:
            #     print(cells)
                # for cell in cells:
                #     rng_values.append(cell.value)
            # dict_item[Name_ColumnTitle] = rng_values
            # print(dict_item)

        # wbMasterFile = load_workbook(filename=".\\Format\\Default_Format.xlsx")
        # rng = ws


if __name__ == "__main__":
    app = App()
    app.mainloop()
