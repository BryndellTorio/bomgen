from openpyxl import load_workbook, Workbook

workbook = load_workbook(filename = "test_data.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(
        min_row=1,
        min_col=1,
        max_col=11,
        values_only=True):
    designator = row[0]
    productNumber = row[9]
    productDescription = row[10]

def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)

print_rows()
print(sheet.dimensions)



# for index in len(designator):
#     ++count
#     print(count)
#     if row[index] == "C1":
#         print("found it!")


# print(designator, productNumber, productDescription)
